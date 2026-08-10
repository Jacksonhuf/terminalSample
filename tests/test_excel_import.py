"""Tests for Excel ontology import."""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from ontology_platform.admin.excel_import import (
    ExcelImportError,
    build_template_bytes,
    import_ontology_from_excel,
    parse_ontology_excel,
)
from ontology_platform.admin.manager import OntologyManager
from ontology_platform.admin.server import create_app
from ontology_platform.ontology.schema import OntologyDef


def _build_workbook_bytes(rows_by_sheet: dict[str, list[list]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "instructions"
    ws.cell(row=1, column=1, value="test")

    for sheet_name, rows in rows_by_sheet.items():
        sheet = wb.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture
def temp_dir(tmp_path):
    return tmp_path


@pytest.fixture
def manager(temp_dir):
    return OntologyManager(temp_dir)


@pytest.fixture
def sample_workbook_bytes():
    return _build_workbook_bytes(
        {
            "ontology": [
                ["name", "version", "description"],
                ["imported_ont", "1.0", "Imported via Excel"],
            ],
            "object_types": [
                ["name", "display_name", "description", "primary_key"],
                ["Asset", "资产", "资产对象", "id"],
                ["Person", "人员", "", "id"],
            ],
            "properties": [
                ["object_type", "name", "type", "required", "enum_values", "description"],
                ["Asset", "id", "string", "true", "", "ID"],
                ["Asset", "status", "enum", "true", "active,inactive", "状态"],
                ["Person", "id", "string", "true", "", "ID"],
            ],
            "links": [
                ["name", "source_type", "target_type", "cardinality", "description"],
                ["owns", "Person", "Asset", "many", "拥有关系"],
            ],
            "actions": [
                [
                    "name",
                    "display_name",
                    "description",
                    "target_type",
                    "requires_approval",
                    "keywords",
                    "allowed_roles",
                    "approver_roles",
                ],
                ["Transfer", "转移", "转移资产", "Asset", "true", "转移", "admin", "admin"],
            ],
            "action_params": [
                ["action_name", "name", "type", "required", "description"],
                ["Transfer", "person_id", "string", "true", "接收人"],
            ],
        }
    )


class TestExcelImport:
    def test_build_template_bytes(self):
        data = build_template_bytes()
        assert data.startswith(b"PK")
        ontology = parse_ontology_excel(data)
        assert ontology.name == "my_ontology"
        assert len(ontology.object_types) == 2
        assert ontology.object_types[0].name == "Asset"
        assert len(ontology.links) == 1
        assert len(ontology.actions) == 1

    def test_parse_workbook(self, sample_workbook_bytes):
        ontology = parse_ontology_excel(sample_workbook_bytes)
        assert ontology.name == "imported_ont"
        assert len(ontology.object_types) == 2
        asset = ontology.get_object_type("Asset")
        assert asset is not None
        assert len(asset.properties) == 2
        assert ontology.get_link("owns") is not None
        action = ontology.get_action("Transfer")
        assert action is not None
        assert action.requires_approval is True
        assert len(action.parameters) == 1

    def test_import_creates_yaml(self, manager, sample_workbook_bytes):
        result = import_ontology_from_excel(sample_workbook_bytes, manager)
        assert result["created"] is True
        loaded = manager.load("imported_ont")
        assert loaded.description == "Imported via Excel"
        assert len(loaded.links) == 1

    def test_import_conflict_without_overwrite(self, manager, sample_workbook_bytes):
        manager.create(OntologyDef(name="imported_ont"))
        with pytest.raises(ExcelImportError, match="已存在"):
            import_ontology_from_excel(sample_workbook_bytes, manager, overwrite=False)

    def test_import_overwrite(self, manager, sample_workbook_bytes):
        manager.create(OntologyDef(name="imported_ont", description="old"))
        result = import_ontology_from_excel(sample_workbook_bytes, manager, overwrite=True)
        assert result["overwritten"] is True
        assert manager.load("imported_ont").description == "Imported via Excel"

    def test_missing_object_type_reference(self, manager):
        data = _build_workbook_bytes(
            {
                "ontology": [["name", "version", "description"], ["bad", "1.0", ""]],
                "object_types": [["name", "display_name", "description", "primary_key"], ["Asset", "", "", "id"]],
                "properties": [],
                "links": [
                    ["name", "source_type", "target_type", "cardinality", "description"],
                    ["bad_link", "Person", "Asset", "many", ""],
                ],
                "actions": [],
                "action_params": [],
            }
        )
        with pytest.raises(ExcelImportError, match="未知对象类型"):
            parse_ontology_excel(data)


class TestExcelImportAPI:
    def test_download_template(self, temp_dir):
        client = TestClient(create_app(temp_dir))
        res = client.get("/api/ontologies/import/template")
        assert res.status_code == 200
        assert "spreadsheetml" in res.headers["content-type"]
        assert res.content.startswith(b"PK")

    def test_import_endpoint(self, temp_dir, sample_workbook_bytes):
        client = TestClient(create_app(temp_dir))
        res = client.post(
            "/api/ontologies/import",
            files={"file": ("ontology.xlsx", sample_workbook_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        data = res.json()
        assert data["ontology"]["name"] == "imported_ont"
        assert OntologyManager(temp_dir).load("imported_ont").name == "imported_ont"

    def test_import_rejects_non_excel(self, temp_dir):
        client = TestClient(create_app(temp_dir))
        res = client.post(
            "/api/ontologies/import",
            files={"file": ("bad.txt", b"hello", "text/plain")},
        )
        assert res.status_code == 400
