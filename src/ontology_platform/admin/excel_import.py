"""Excel template generation and ontology import."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ontology_platform.admin.manager import OntologyManager
from ontology_platform.ontology.schema import (
    ActionDef,
    ActionParamDef,
    LinkDef,
    ObjectTypeDef,
    OntologyDef,
    PropertyDef,
)

SHEET_ONTOLOGY = "ontology"
SHEET_OBJECT_TYPES = "object_types"
SHEET_PROPERTIES = "properties"
SHEET_LINKS = "links"
SHEET_ACTIONS = "actions"
SHEET_ACTION_PARAMS = "action_params"
SHEET_INSTRUCTIONS = "instructions"

ONTOLOGY_HEADERS = ["name", "version", "description"]
OBJECT_TYPE_HEADERS = ["name", "display_name", "description", "primary_key"]
PROPERTY_HEADERS = ["object_type", "name", "type", "required", "enum_values", "description"]
LINK_HEADERS = ["name", "source_type", "target_type", "cardinality", "description"]
ACTION_HEADERS = [
    "name",
    "display_name",
    "description",
    "target_type",
    "requires_approval",
    "keywords",
    "allowed_roles",
    "approver_roles",
]
ACTION_PARAM_HEADERS = ["action_name", "name", "type", "required", "description"]

TEMPLATE_ROWS = {
    SHEET_ONTOLOGY: [["my_ontology", "1.0", "本体描述"]],
    SHEET_OBJECT_TYPES: [
        ["Asset", "资产", "资产对象", "id"],
        ["Person", "人员", "组织成员", "id"],
    ],
    SHEET_PROPERTIES: [
        ["Asset", "id", "string", "true", "", "唯一标识"],
        ["Asset", "name", "string", "true", "", "名称"],
        ["Asset", "status", "enum", "true", "active,inactive", "状态"],
        ["Person", "id", "string", "true", "", "唯一标识"],
        ["Person", "name", "string", "true", "", "姓名"],
    ],
    SHEET_LINKS: [["owns", "Person", "Asset", "many", "人员拥有资产"]],
    SHEET_ACTIONS: [
        [
            "TransferAsset",
            "转移资产",
            "将资产转移给另一人员",
            "Asset",
            "true",
            "转移,transfer",
            "operator,admin",
            "admin",
        ]
    ],
    SHEET_ACTION_PARAMS: [
        ["TransferAsset", "person_id", "string", "true", "接收人 ID"],
    ],
}

INSTRUCTION_LINES = [
    "本体 Excel 导入说明",
    "",
    "1. ontology 工作表填写一行本体元数据（name 为英文标识，必填）",
    "2. object_types 填写对象类型；properties 通过 object_type 列关联对象类型",
    "3. links 的 source_type / target_type 必须引用已定义的对象类型",
    "4. actions 的 target_type 必须引用已定义的对象类型；参数在 action_params 中填写",
    "5. required / requires_approval 支持: true/false、yes/no、1/0、是/否",
    "6. enum_values、keywords、allowed_roles、approver_roles 使用英文逗号分隔",
    "7. 导入时若本体已存在，可勾选「覆盖已有本体」",
]


class ExcelImportError(ValueError):
    """Raised when Excel content cannot be parsed into an ontology."""


def build_template_bytes() -> bytes:
    """Return an Excel workbook template as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_INSTRUCTIONS
    for idx, line in enumerate(INSTRUCTION_LINES, start=1):
        ws.cell(row=idx, column=1, value=line)

    for sheet_name, headers in [
        (SHEET_ONTOLOGY, ONTOLOGY_HEADERS),
        (SHEET_OBJECT_TYPES, OBJECT_TYPE_HEADERS),
        (SHEET_PROPERTIES, PROPERTY_HEADERS),
        (SHEET_LINKS, LINK_HEADERS),
        (SHEET_ACTIONS, ACTION_HEADERS),
        (SHEET_ACTION_PARAMS, ACTION_PARAM_HEADERS),
    ]:
        sheet = wb.create_sheet(sheet_name)
        _write_header(sheet, headers)
        for row in TEMPLATE_ROWS.get(sheet_name, []):
            sheet.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_ontology_excel(data: bytes) -> OntologyDef:
    """Parse an uploaded Excel workbook into an OntologyDef."""
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportError(f"无法读取 Excel 文件: {exc}") from exc

    try:
        ontology_meta = _read_single_row(wb, SHEET_ONTOLOGY, ONTOLOGY_HEADERS)
        name = _require_str(ontology_meta, "name", SHEET_ONTOLOGY)
        version = _optional_str(ontology_meta.get("version")) or "1.0"
        description = _optional_str(ontology_meta.get("description")) or ""

        object_types = _parse_object_types(wb)
        type_names = {ot.name for ot in object_types}

        links = _parse_links(wb, type_names)
        actions = _parse_actions(wb, type_names)

        return OntologyDef(
            name=name,
            version=version,
            description=description,
            object_types=object_types,
            links=links,
            actions=actions,
        )
    finally:
        wb.close()


def import_ontology_from_excel(
    data: bytes,
    manager: OntologyManager,
    *,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Parse Excel and persist ontology YAML via OntologyManager."""
    ontology = parse_ontology_excel(data)
    exists = _ontology_exists(manager, ontology.name)

    if exists and not overwrite:
        raise ExcelImportError(f"本体已存在: {ontology.name}，请勾选覆盖或更换名称")

    path = manager.save(ontology)
    return {
        "message": "imported",
        "created": not exists,
        "overwritten": exists and overwrite,
        "path": str(path),
        "ontology": ontology.model_dump(),
        "stats": {
            "object_types": len(ontology.object_types),
            "links": len(ontology.links),
            "actions": len(ontology.actions),
        },
    }


def _ontology_exists(manager: OntologyManager, name: str) -> bool:
    try:
        manager.load(name)
        return True
    except FileNotFoundError:
        return False


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)


def _read_single_row(wb, sheet_name: str, expected_headers: list[str]) -> dict[str, Any]:
    rows = _read_sheet_rows(wb, sheet_name, expected_headers)
    if not rows:
        raise ExcelImportError(f"工作表 {sheet_name} 至少需要一行数据")
    if len(rows) > 1:
        raise ExcelImportError(f"工作表 {sheet_name} 只能填写一行本体元数据")
    return rows[0]


def _read_sheet_rows(wb, sheet_name: str, expected_headers: list[str]) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = [_normalize_header(cell) for cell in rows[0]]
    if not any(header_row):
        return []

    missing = [h for h in expected_headers if h not in header_row]
    if missing:
        raise ExcelImportError(f"工作表 {sheet_name} 缺少列: {', '.join(missing)}")

    parsed: list[dict[str, Any]] = []
    for row in rows[1:]:
        if _is_empty_row(row):
            continue
        item = {header_row[i]: row[i] if i < len(row) else None for i in range(len(header_row)) if header_row[i]}
        parsed.append(item)
    return parsed


def _parse_object_types(wb) -> list[ObjectTypeDef]:
    type_rows = _read_sheet_rows(wb, SHEET_OBJECT_TYPES, OBJECT_TYPE_HEADERS)
    if not type_rows:
        raise ExcelImportError("至少需要定义一个对象类型 (object_types)")

    properties_by_type: dict[str, list[PropertyDef]] = {}
    for row in _read_sheet_rows(wb, SHEET_PROPERTIES, PROPERTY_HEADERS):
        object_type = _require_str(row, "object_type", SHEET_PROPERTIES)
        prop_name = _require_str(row, "name", SHEET_PROPERTIES)
        prop_type = _optional_str(row.get("type")) or "string"
        properties_by_type.setdefault(object_type, []).append(
            PropertyDef(
                name=prop_name,
                type=prop_type,
                required=_parse_bool(row.get("required"), default=False),
                enum_values=_parse_list(row.get("enum_values")),
                description=_optional_str(row.get("description")) or "",
            )
        )

    object_types: list[ObjectTypeDef] = []
    seen: set[str] = set()
    for row in type_rows:
        name = _require_str(row, "name", SHEET_OBJECT_TYPES)
        if name in seen:
            raise ExcelImportError(f"对象类型重复: {name}")
        seen.add(name)
        object_types.append(
            ObjectTypeDef(
                name=name,
                display_name=_optional_str(row.get("display_name")) or "",
                description=_optional_str(row.get("description")) or "",
                primary_key=_optional_str(row.get("primary_key")) or "id",
                properties=properties_by_type.get(name, []),
            )
        )

    undefined_props = sorted(set(properties_by_type) - seen)
    if undefined_props:
        raise ExcelImportError(f"properties 引用了未定义的对象类型: {', '.join(undefined_props)}")

    return object_types


def _parse_links(wb, type_names: set[str]) -> list[LinkDef]:
    links: list[LinkDef] = []
    seen: set[str] = set()
    for row in _read_sheet_rows(wb, SHEET_LINKS, LINK_HEADERS):
        name = _require_str(row, "name", SHEET_LINKS)
        if name in seen:
            raise ExcelImportError(f"关系重复: {name}")
        seen.add(name)
        source_type = _require_str(row, "source_type", SHEET_LINKS)
        target_type = _require_str(row, "target_type", SHEET_LINKS)
        _ensure_type_exists(source_type, type_names, SHEET_LINKS, "source_type")
        _ensure_type_exists(target_type, type_names, SHEET_LINKS, "target_type")
        links.append(
            LinkDef(
                name=name,
                source_type=source_type,
                target_type=target_type,
                cardinality=_optional_str(row.get("cardinality")) or "many",
                description=_optional_str(row.get("description")) or "",
            )
        )
    return links


def _parse_actions(wb, type_names: set[str]) -> list[ActionDef]:
    params_by_action: dict[str, list[ActionParamDef]] = {}
    for row in _read_sheet_rows(wb, SHEET_ACTION_PARAMS, ACTION_PARAM_HEADERS):
        action_name = _require_str(row, "action_name", SHEET_ACTION_PARAMS)
        param_name = _require_str(row, "name", SHEET_ACTION_PARAMS)
        params_by_action.setdefault(action_name, []).append(
            ActionParamDef(
                name=param_name,
                type=_optional_str(row.get("type")) or "string",
                required=_parse_bool(row.get("required"), default=True),
                description=_optional_str(row.get("description")) or "",
            )
        )

    actions: list[ActionDef] = []
    seen: set[str] = set()
    for row in _read_sheet_rows(wb, SHEET_ACTIONS, ACTION_HEADERS):
        name = _require_str(row, "name", SHEET_ACTIONS)
        if name in seen:
            raise ExcelImportError(f"动作重复: {name}")
        seen.add(name)
        target_type = _require_str(row, "target_type", SHEET_ACTIONS)
        _ensure_type_exists(target_type, type_names, SHEET_ACTIONS, "target_type")
        actions.append(
            ActionDef(
                name=name,
                display_name=_optional_str(row.get("display_name")) or "",
                description=_optional_str(row.get("description")) or "",
                target_type=target_type,
                requires_approval=_parse_bool(row.get("requires_approval"), default=False),
                keywords=_parse_list(row.get("keywords")),
                allowed_roles=_parse_list(row.get("allowed_roles")),
                approver_roles=_parse_list(row.get("approver_roles")),
                parameters=params_by_action.get(name, []),
            )
        )

    undefined_params = sorted(set(params_by_action) - seen)
    if undefined_params:
        raise ExcelImportError(f"action_params 引用了未定义的动作: {', '.join(undefined_params)}")

    return actions


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _require_str(row: dict[str, Any], key: str, sheet: str) -> str:
    value = _optional_str(row.get(key))
    if not value:
        raise ExcelImportError(f"工作表 {sheet} 的 {key} 不能为空")
    return value


def _ensure_type_exists(type_name: str, type_names: set[str], sheet: str, column: str) -> None:
    if type_name not in type_names:
        raise ExcelImportError(f"工作表 {sheet} 的 {column} 引用了未知对象类型: {type_name}")


def _parse_bool(value: Any, *, default: bool) -> bool:
    if value is None or str(value).strip() == "":
        return default
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y", "是", "真"}:
        return True
    if text in {"false", "0", "no", "n", "否", "假"}:
        return False
    raise ExcelImportError(f"无法解析布尔值: {value}")


def _parse_list(value: Any) -> list[str]:
    text = _optional_str(value)
    if not text:
        return []
    return [part.strip() for part in text.replace("，", ",").split(",") if part.strip()]
